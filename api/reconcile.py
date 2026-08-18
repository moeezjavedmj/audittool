"""
Real Python backend for the reconciliation engine, deployed as a native
Vercel Python Function (BaseHTTPRequestHandler entrypoint - Vercel's
zero-config Python runtime).

This does NOT run on Netlify - Netlify Functions only execute JS/TS, Go,
or Rust, so a .py file here is simply inert on that platform. The frontend
calls this endpoint first and transparently falls back to the Node engine
(/api/process, identical logic, works everywhere) when this route isn't
available - see the fetch logic in app/page.tsx.

Files are sent as base64 JSON rather than multipart/form-data, since
Vercel's Python runtime has no bundled multipart parser and this keeps
the handler pure standard library plus openpyxl.

Request body:
{
  "mode": "active" | "resigned",
  "file1": {"name": "hr.xlsx", "dataBase64": "..."},
  "file2": {"name": "dept.xlsx", "dataBase64": "..."}
}
"""
import base64
import io
import json
from http.server import BaseHTTPRequestHandler

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- Normalization helpers (same rules as the Node/TS and original Python
# versions, so all three engines agree on every dataset) ---------------------

def normalize_id(value):
    if value is None:
        return ""
    text = str(value)
    if text.strip().lower() in ("nan", "none", ""):
        return ""
    text = " ".join(text.strip().split())
    return text.upper()


ID_COLUMN_CANDIDATES = [
    "employee id", "employeeid", "emp id", "empid",
    "employee number", "employee no", "emp no",
    "staff id", "staff no", "personnel id", "personnel number",
    "user id", "userid", "id"
]


def guess_id_column(columns):
    norm_map = {" ".join(str(c).strip().lower().split()): c for c in columns}
    for cand in ID_COLUMN_CANDIDATES:
        if cand in norm_map:
            return norm_map[cand]
    for norm, orig in norm_map.items():
        if "employee" in norm and "id" in norm:
            return orig
    for norm, orig in norm_map.items():
        if norm.endswith("id"):
            return orig
    return columns[0]


STATUS_COLUMN_CANDIDATES = [
    "status", "employee status", "emp status", "work status",
    "employment status", "active status", "account status", "state"
]

ACTIVE_VALUES = {
    "A", "ACTIVE", "ACTIVE EMPLOYEE", "ACTIVELY WORKING", "WORKING",
    "CURRENT", "CURRENTLY WORKING", "ENABLED", "ENABLE", "Y", "YES", "1",
    "TRUE", "IN SERVICE", "EMPLOYED", "ON DUTY", "PRESENT"
}

INACTIVE_VALUES = {
    "I", "INACTIVE", "RESIGNED", "TERMINATED", "TERMINATE", "LEFT",
    "EXIT", "EXITED", "SEPARATED", "N", "NO", "0", "FALSE", "DISABLED",
    "DISABLE", "SUSPENDED", "OFFBOARDED", "NOT ACTIVE"
}


def guess_status_column(columns):
    norm_map = {" ".join(str(c).strip().lower().split()): c for c in columns}
    for cand in STATUS_COLUMN_CANDIDATES:
        if cand in norm_map:
            return norm_map[cand]
    for norm, orig in norm_map.items():
        if "status" in norm:
            return orig
    return None


def classify_status(raw_value):
    norm = normalize_id(raw_value)
    if norm in ACTIVE_VALUES:
        return "active"
    if norm in INACTIVE_VALUES:
        return "inactive"
    return "unknown"


def guess_name_column(columns):
    candidates = ["employee name", "name", "full name", "staff name"]
    norm_map = {" ".join(str(c).strip().lower().split()): c for c in columns}
    for cand in candidates:
        if cand in norm_map:
            return norm_map[cand]
    for norm, orig in norm_map.items():
        if "name" in norm:
            return orig
    return None


# ---- Format detection + reading ---------------------------------------------

def sniff_format(data: bytes):
    sig = data[:8]
    if sig[:2] == b"PK":
        return "xlsx"
    if sig[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":
        return "xls"
    return "csv"


def load_table(data: bytes, declared_name: str):
    declared_ext = ""
    if "." in declared_name:
        declared_ext = "." + declared_name.rsplit(".", 1)[-1].lower()
    actual = sniff_format(data)

    if actual == "xls":
        raise ValueError(
            "This file is a legacy .xls (old binary Excel) file, which isn't "
            "supported. Please re-save it as .xlsx and upload that instead."
        )
    if actual == "csv" and declared_ext in (".xlsx", ".xlsm"):
        raise ValueError(
            f"This file is named like an Excel file ({declared_ext}) but its "
            "actual content is plain text/CSV, not a real workbook."
        )

    if actual == "xlsx":
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows_iter = ws.iter_rows(values_only=True)
        try:
            header = next(rows_iter)
        except StopIteration:
            wb.close()
            return [], []
        columns = [str(h) if h is not None else f"Column{i+1}" for i, h in enumerate(header)]
        raw_rows = []
        for r in rows_iter:
            if r is None or all(v is None or str(v).strip() == "" for v in r):
                continue
            raw_rows.append(list(r))
        wb.close()
    else:
        text = data.decode("utf-8-sig", errors="strict")
        import csv as csv_module
        reader = csv_module.reader(io.StringIO(text))
        rows_list = list(reader)
        if not rows_list:
            return [], []
        columns = rows_list[0]
        raw_rows = [row for row in rows_list[1:] if any(str(v).strip() != "" for v in row)]

    keep_idx = []
    for i in range(len(columns)):
        has_value = any(
            i < len(r) and r[i] is not None and str(r[i]).strip() != "" for r in raw_rows
        )
        if has_value or (columns[i] and str(columns[i]).strip() != ""):
            keep_idx.append(i)

    columns = [str(columns[i]).strip() for i in keep_idx]
    seen = {}
    deduped = []
    for c in columns:
        if c not in seen:
            seen[c] = 0
            deduped.append(c)
        else:
            seen[c] += 1
            deduped.append(f"{c} ({seen[c]})")
    columns = deduped

    rows = []
    for r in raw_rows:
        row_dict = {}
        for pos, i in enumerate(keep_idx):
            val = r[i] if i < len(r) else None
            row_dict[columns[pos]] = "" if val is None else str(val)
        rows.append(row_dict)

    return columns, rows


# ---- Writing the styled report ----------------------------------------------

HEADER_FILL = PatternFill(start_color="0D1117", end_color="0D1117", fill_type="solid")
HEADER_FONT = Font(color="3FD9DB", bold=True, size=11, name="Calibri")
FLAG_FILL = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid")
THIN = Side(style="thin", color="D0D5DD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def write_report_bytes(columns, rows, sheet_title):
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title[:31]

    ws.append(columns)
    for col_idx in range(1, len(columns) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left", vertical="center")
        cell.border = BORDER

    for row in rows:
        ws.append([row.get(c, "") for c in columns])

    for r in range(2, len(rows) + 2):
        for c in range(1, len(columns) + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = BORDER
            cell.fill = FLAG_FILL

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = len(str(col_name))
        for row in rows:
            v = row.get(col_name, "")
            if len(str(v)) > max_len:
                max_len = len(str(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 45)

    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


# ---- Modes -------------------------------------------------------------------

def run_active_mode(hr_bytes, hr_name, dept_bytes, dept_name):
    hr_columns, hr_rows = load_table(hr_bytes, hr_name)
    dept_columns, dept_rows = load_table(dept_bytes, dept_name)

    hr_id_col = guess_id_column(hr_columns)
    dept_id_col = guess_id_column(dept_columns)
    status_col = guess_status_column(hr_columns)

    considered_hr_rows = hr_rows
    active_only_applied = False
    if status_col is not None and hr_rows:
        classes = [classify_status(r.get(status_col, "")) for r in hr_rows]
        known = [c in ("active", "inactive") for c in classes]
        if sum(known) / len(known) > 0.5:
            considered_hr_rows = [r for r, c in zip(hr_rows, classes) if c == "active"]
            active_only_applied = True

    hr_id_set = {normalize_id(r.get(hr_id_col, "")) for r in considered_hr_rows}
    hr_id_set.discard("")

    missing_rows = []
    matched_count = 0
    for r in dept_rows:
        did = normalize_id(r.get(dept_id_col, ""))
        if did == "":
            continue
        if did in hr_id_set:
            matched_count += 1
        else:
            missing_rows.append(r)

    report_bytes = write_report_bytes(dept_columns, missing_rows, "Not In Active List")
    name_col = guess_name_column(dept_columns)

    return {
        "mode": "active",
        "hrIdColumn": hr_id_col,
        "deptIdColumn": dept_id_col,
        "statusColumn": status_col,
        "activeFilterApplied": active_only_applied,
        "hrRows": len(hr_rows),
        "hrActiveRows": len(considered_hr_rows),
        "deptRows": len(dept_rows),
        "matched": matched_count,
        "flagged": len(missing_rows),
        "flagLabel": "Dept users missing from active list",
        "nameColumn": name_col,
        "columns": dept_columns,
        "preview": missing_rows[:50],
        "reportBase64": base64.b64encode(report_bytes).decode("ascii"),
        "reportFileName": "reconciliation-report.xlsx",
        "engine": "python",
    }


def run_resigned_mode(resigned_bytes, resigned_name, dept_bytes, dept_name):
    resigned_columns, resigned_rows = load_table(resigned_bytes, resigned_name)
    dept_columns, dept_rows = load_table(dept_bytes, dept_name)

    resigned_id_col = guess_id_column(resigned_columns)
    dept_id_col = guess_id_column(dept_columns)

    resigned_id_set = {normalize_id(r.get(resigned_id_col, "")) for r in resigned_rows}
    resigned_id_set.discard("")

    still_present_rows = []
    for r in dept_rows:
        did = normalize_id(r.get(dept_id_col, ""))
        if did != "" and did in resigned_id_set:
            still_present_rows.append(r)

    report_bytes = write_report_bytes(dept_columns, still_present_rows, "Resigned But Still Listed")
    name_col = guess_name_column(dept_columns)

    return {
        "mode": "resigned",
        "resignedIdColumn": resigned_id_col,
        "deptIdColumn": dept_id_col,
        "resignedRows": len(resigned_rows),
        "deptRows": len(dept_rows),
        "flagged": len(still_present_rows),
        "flagLabel": "Resigned employees still listed in department",
        "nameColumn": name_col,
        "columns": dept_columns,
        "preview": still_present_rows[:50],
        "reportBase64": base64.b64encode(report_bytes).decode("ascii"),
        "reportFileName": "reconciliation-report.xlsx",
        "engine": "python",
    }


# ---- Vercel Python Function entrypoint --------------------------------------

class handler(BaseHTTPRequestHandler):
    def _send_json(self, status, payload):
        body = json.dumps(payload).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json")
        self.send_header("Content-Length", str(len(body)))
        self.end_headers()
        self.wfile.write(body)

    def do_POST(self):
        try:
            length = int(self.headers.get("Content-Length", 0))
            raw = self.rfile.read(length)
            payload = json.loads(raw.decode("utf-8"))

            mode = payload.get("mode")
            file1 = payload.get("file1") or {}
            file2 = payload.get("file2") or {}

            if mode not in ("active", "resigned"):
                self._send_json(400, {"ok": False, "error": "Invalid mode."})
                return
            if not file1.get("dataBase64") or not file2.get("dataBase64"):
                self._send_json(400, {"ok": False, "error": "Both files are required."})
                return

            file1_bytes = base64.b64decode(file1["dataBase64"])
            file2_bytes = base64.b64decode(file2["dataBase64"])

            if mode == "active":
                result = run_active_mode(file1_bytes, file1.get("name", ""), file2_bytes, file2.get("name", ""))
            else:
                result = run_resigned_mode(file1_bytes, file1.get("name", ""), file2_bytes, file2.get("name", ""))

            result["ok"] = True
            self._send_json(200, result)
        except Exception as exc:  # noqa: BLE001
            self._send_json(500, {"ok": False, "error": str(exc)})
